"""
Module de chargement et préparation de la base médicaments.
Fournit :
  - build_whisper_prompt()        → prompt court pour Whisper API Groq (≤ 900 chars)
  - build_whisper_prompt_full()   → prompt complet pour Whisper local (aucune limite)
  - get_meds_context()            → contexte complet pour Groq (correction / standardisation)
  - lookup(nom)                   → cherche un médicament par nom ou DCI
  - find_meds_in_text(text)       → détecte les médicaments dans un texte
"""

import os
import json
import re
from functools import lru_cache
from openpyxl import load_workbook

XLSX_PATH = os.path.join(os.path.dirname(__file__), "../models/medicaments_final_amir.xlsx")


@lru_cache(maxsize=1)
def load_meds() -> dict:
    """Charge le fichier Excel une seule fois en mémoire."""
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    noms = set()
    dcis = set()
    entries = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        nom    = str(row[1]).strip() if row[1] else ""
        dci    = str(row[2]).strip() if row[2] else ""
        valeur = str(row[10]).strip() if row[10] else ""
        unite  = str(row[11]).strip() if row[11] else ""
        forme  = str(row[7]).strip() if row[7] else ""

        if nom: noms.add(nom)
        if dci: dcis.add(dci)
        if nom:
            entries.append({
                "nom": nom,
                "dci": dci,
                "dosage": f"{valeur} {unite}".strip(),
                "forme": forme,
            })

    return {
        "noms": sorted(noms),
        "dcis": sorted(dcis),
        "entries": entries,
    }


def build_whisper_prompt() -> str:
    """
    Prompt court pour Whisper via API Groq (≤ 900 caractères ≈ 224 tokens).
    Priorité aux DCI courtes, puis noms commerciaux.
    """
    data = load_meds()
    noms_sorted = sorted(data["noms"], key=len)
    dcis_sorted = sorted(data["dcis"], key=len)

    budget = 900
    intro = "Note médicale. Médicaments et DCI: "
    parts = [intro]
    used = len(intro)

    for term in dcis_sorted + noms_sorted:
        chunk = term + ", "
        if used + len(chunk) > budget:
            break
        parts.append(chunk)
        used += len(chunk)

    return "".join(parts).rstrip(", ") + "."


def build_whisper_prompt_full() -> str:
    """
    Prompt COMPLET pour Whisper local (openai-whisper) — aucune limite de taille.

    Stratégie :
    - Phrase d'intro médicale pour ancrer Whisper dans le bon domaine/langue.
    - Toutes les DCI en premier (prononciation générique, plus spoken).
    - Tous les noms commerciaux ensuite.
    - Les termes sont dédupliqués (insensible à la casse).

    Whisper utilise ce texte comme "previous segment" interne :
    il biaise le décodage beam search vers ces tokens, améliorant
    significativement la reconnaissance des termes rares.
    """
    data = load_meds()

    # DCIs en premier (généralement la prononciation orale), puis noms commerciaux
    dcis_sorted = sorted(data["dcis"], key=str.lower)
    noms_sorted = sorted(data["noms"], key=str.lower)

    # Dédupliquer (un nom commercial = parfois identique à la DCI)
    seen = set()
    terms = []
    for t in dcis_sorted + noms_sorted:
        if t.lower() not in seen:
            seen.add(t.lower())
            terms.append(t)

    intro = (
        "Note médicale rédigée par un délégué médical francophone. "
        "Médicaments, DCI et termes médicaux mentionnés : "
    )
    return intro + ", ".join(terms) + "."


def get_meds_context(max_entries: int = 200) -> str:
    """
    Retourne un bloc JSON compact des médicaments pour le contexte Groq.
    Limité à max_entries pour ne pas dépasser la fenêtre de contexte.
    """
    data = load_meds()
    sample = data["entries"][:max_entries]
    return json.dumps(sample, ensure_ascii=False)


def lookup(query: str) -> list:
    """Cherche un médicament par nom commercial ou DCI (insensible à la casse)."""
    data = load_meds()
    q = query.strip().lower()
    return [
        e for e in data["entries"]
        if q in e["nom"].lower() or q in e["dci"].lower()
    ]


def find_meds_in_text(text: str) -> list:
    """Détecte les noms de médicaments présents dans un texte."""
    data = load_meds()
    text_upper = text.upper()
    found = []
    for nom in data["noms"]:
        if re.search(r'\b' + re.escape(nom) + r'\b', text_upper):
            found.append(nom)
    for dci in data["dcis"]:
        if re.search(r'\b' + re.escape(dci) + r'\b', text_upper):
            found.append(dci)
    return list(set(found))